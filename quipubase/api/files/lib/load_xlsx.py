import base64
import json
import typing as tp
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal

import typing_extensions as tpe
from openpyxl import load_workbook

from ._base import Artifact


class JsonEncoder(json.JSONEncoder):
    def default(self, o: object) -> object:
        if isinstance(o, datetime):
            return o.isoformat()
        if isinstance(o, date):
            return o.isoformat()
        if isinstance(o, time):
            return o.isoformat()
        if isinstance(o, timedelta):
            return o.total_seconds()
        if isinstance(o, Decimal):
            return str(o)
        return super().default(o)


@dataclass
class ExcelLoader(Artifact):
    ref: tpe.Annotated[
        str,
        tpe.Doc(
            """
        This `ref` can represent one out of three things:

        - An HTTP URL.
        - A file path (temporary or not) within the local filesystem.
        - A text file content.
        """
        ),
    ]

    def extract(self) -> tp.Generator[str, None, None]:
        """
        Extract data from Excel spreadsheets.

        Yields:
            JSON strings representing cell data with sheet, position, and value
        """
        file_path = self.retrieve()

        try:
            # Load the Excel workbook with data_only=True to get values, not formulas
            wb = load_workbook(filename=file_path, data_only=True)

            # Process each sheet in the workbook
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Add sheet metadata
                yield json.dumps(
                    {
                        "type": "sheet_metadata",
                        "sheet": sheet_name,
                        "rows": sheet.max_row,
                        "columns": sheet.max_column,
                    },
                    cls=JsonEncoder,
                )

                # Extract data from each cell
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is not None:  # Only process cells with values
                            try:
                                data_dict = {  # type: ignore
                                    "sheet": sheet_name,
                                    "position": f"{cell.column_letter}{cell.row}",  # type: ignore
                                    "value": cell.value,
                                }
                                yield json.dumps(data_dict, cls=JsonEncoder)
                            except (TypeError, ValueError) as e:
                                # Handle cells with values that can't be serialized
                                yield json.dumps(
                                    {
                                        "sheet": sheet_name,
                                        "position": f"{cell.column_letter}{cell.row}",  # type: ignore
                                        "value": str(cell.value),
                                        "error": str(e),
                                    },
                                    cls=JsonEncoder,
                                )
        except Exception as e:
            # Handle errors gracefully
            yield json.dumps(
                {"type": "error", "message": f"Error processing Excel file: {str(e)}"},
                cls=JsonEncoder,
            )
